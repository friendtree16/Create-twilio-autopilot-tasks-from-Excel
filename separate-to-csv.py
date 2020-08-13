# coding:utf-8
import openpyxl
import sys
import os
import shutil
import argparse
import json
import codecs

if len(sys.argv) == 1:
    print('plese input file name.')
    sys.exit()
fileName = sys.argv[1]
book = openpyxl.load_workbook(fileName)
targetSheet = book['tasks']

directoryName = 'upload/'
questionDirectoryName = 'questions/'
answerDirectoryName = 'answer/'
if os.path.exists(directoryName) == True:
    shutil.rmtree(directoryName)
os.makedirs(directoryName)

maxRow = targetSheet.max_row + 1
# ヘッダー行を指定
firstRow = 2
taskName = ''

# インポート用のQuestionファイル作成
os.makedirs(directoryName + questionDirectoryName)
for count in range(firstRow, maxRow):
    # インテント名取得
    intent = targetSheet.cell(row=count, column=4).value
    if intent != None:
        if taskName != '':
            file.close()
        taskName = intent
        csvFileName = directoryName + questionDirectoryName + taskName + '.csv'
        file = open(csvFileName, 'a')
            
    # 質問内容取得
    question = targetSheet.cell(row=count, column=3).value
    question += '\n'
    file.write(question.encode('utf-8'))

    
 
file.close()

# 回答取得
# インポート用Answer
os.makedirs(directoryName + answerDirectoryName)
for count in range(firstRow, maxRow):
    # インテント名取得
    intent = targetSheet.cell(row=count, column=4).value
    if intent != None:
        taskName = intent
        csvFileName = directoryName + answerDirectoryName + taskName + '.json'

        answer = targetSheet.cell(row=count, column=6).value

        answerJson = {
            'actions': [
                {
                    'say': answer
                }
            ]
        }

        file = codecs.open(csvFileName, 'w', 'utf-8')
        json.dump(answerJson, file ,ensure_ascii=False)
        file.close()
